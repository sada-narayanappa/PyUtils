import numpy as np
from sklearn import preprocessing
from sklearn.decomposition import PCA
import pandas as pd
import os
from IPython.display import display
from IPython.display import HTML
import dateutil;
import json;
import urllib.request;
import re
import base64
from pandas import ExcelFile
from io import StringIO, BytesIO
import xml.etree.ElementTree as ET
import json, requests

import matplotlib
#matplotlib.style.use('ggplot')

np.set_printoptions(precision=2, linewidth=100)
pd.set_option('display.width', 1000)
pd.set_option('precision',5);
pd.set_option('display.precision', 5)
pd.set_option('mode.sim_interactive', True);
pd.set_option('display.max_rows', 9)
pd.set_option('display.float_format', lambda x: '%.3f' % x)

'''
Loads data set for person comparative analysis
Problems and ISSUES - Check the following:
==========================================
** Does your data have headers! If not you need more complex call
If so, Pass "headers=None" and pass names

'''

#Make this generic
class Map(dict):
    """
    Example:
    m = Map({'first_name': 'Eduardo'}, last_name='Pool', age=24, sports=['Soccer'])
    """
    def __init__(self, *args, **kwargs):
        super(Map, self).__init__(*args, **kwargs)
        for arg in args:
            if isinstance(arg, dict):
                for k, v in arg.items():
                    self[k] = v

        if kwargs:
            for k, v in kwargs.iteritems():
                self[k] = v

    def __getattr__(self, attr):
        return self.get(attr)

    def __setattr__(self, key, value):
        self.__setitem__(key, value)

    def __setitem__(self, key, value):
        super(Map, self).__setitem__(key, value)
        self.__dict__.update({key: value})

    def __delattr__(self, item):
        self.__delitem__(item)

    def __delitem__(self, key):
        super(Map, self).__delitem__(key)
        del self.__dict__[key]
        
# LIMITATIONS: Works only if all nodes have same tags
# 1M rows must be good enough
#
def getDFFromXML(file, xmlTag=None, maxRows=1000000):
    with open(file, "r") as f:
        xmlText = f.read()

    root = ET.fromstring(xmlText)
    
    if (xmlTag):
        iters= root.findall(path=xmlTag)
    else:
        iters = root;
    
    cols=[]
    for i,e in enumerate(iters):
        for j,e1 in enumerate(e):
            cols.append(e1.tag)
        break;
    
    rr=[]
    for i,e in enumerate(iters):
        r = []
        for j,e1 in enumerate(e):
             r.append(e1.text)
        rr.append(r)    
        if (i>maxRows):
            break;

    df=pd.DataFrame(rr)
    df.columns = cols
    
    return df;

def DetermineSeperator(line):
    sep = ","    
    split2  = line.split("\t");
    if (len(split2) > 1 ):
        sep = "\t";   
    return sep;
    
def getAuraDF(link, proxies=None):
    #f = urllib.request.urlopen(link)
    #js = f.read().decode('UTF-8')
    f = requests.get(link, verify= False, proxies=proxies)
    js = f.text;
    fjs=re.sub('[\n\r|\r\n|\n\s*]+', '\n', js)
    js=re.sub('^\n', '', fjs)
    if (js.find("$rs=") > 0):
        js = js[js.find("$rs=")+5:]
        #print(js[0:100])
        data = json.loads(js)
        df=pd.DataFrame(data['rows'],columns=data['colnames'])
        return df
    else:
        return js

def getFormulas(ws, df2):
    from collections import defaultdict
    formulas= defaultdict(str);

    rows = df2.shape[0]
    cols = df2.shape[1]    
    for i in range(rows):
        for j in range(cols):
            ii = i+1
            jj = j+1
                
            c= ws.cell(row=ii, column=jj);
            v = c.value;
            
            if ( v is not None and type(v) == str and v.strip().startswith('=') ):
                formulas[(ii,jj)] = v
    return formulas

def getExcelFile(fileName, sheetname=0):
    import openpyxl
    df2 = pd.read_excel(fileName, header=None, sheetname=sheetname)
    colNames = [openpyxl.utils.get_column_letter(c) for c in range(1,df2.shape[1]+1) ]
    df2.columns = (colNames)

    df2.fillna('', inplace=True)
    df2.index = range(len(df2))
    wb = openpyxl.load_workbook(fileName)
    if ( type(sheetname) == int):
        sheetname = wb.sheetnames[sheetname]
    ws=wb[sheetname]
    df2.formulas = getFormulas(ws, df2)
    
    return df2;
    
def getDF(fileName, debug=False, headers=None, names=None, usecols=None, checkForDateTime=False, 
          seperator=None, index_col=None, sheetname=0, xmlTag=None, proxies=None):
    
    if (    not (fileName.startswith("http://"))  and
            not (fileName.startswith("https://")) and
            not os.path.exists(fileName)):
        #raise Exception( fileName + " does not exist")
        print ("ERROR: *** " +fileName + " does not exist");
        return None;
    
    sep = seperator or ",";
    df1=None
    if (fileName.startswith("http")):
        df1 = getAuraDF(fileName, proxies=proxies)
    elif fileName.endswith(".xlsx") or fileName.endswith(".xlsm") or fileName.endswith(".xlsb"): 
        df1 = getExcelFile(fileName, sheetname)
    elif (fileName.endswith(".xml")):
        df1 = getDFFromXML(fileName, xmlTag)
    elif ("/aura/" in fileName):
        df1 = getAuraDF(fileName);
        return df1
    else:
        sep = ","
        if not fileName.endswith(".csv"):
            with open(fileName, 'r') as f:
                line    = f.readline();    
                #split1  = line.split(",");
                sep = DetermineSeperator(line);
                
        df1 = pd.read_csv(fileName, sep=sep, header=headers, low_memory=False,
                       skipinitialspace =True, names=names, comment='#', usecols=usecols)
    return df1;

    
def LoadDataSet(fileOrString, columns=None, excel = False,
                debug=False, headers=0, names=None, checkForDateTime=False, usecols=None,
                seperator=None, index_col=None,sheetname=0, xmlTag=None, proxies=None, **kwargs):
    
    if(type(fileOrString) == bytes and excel ):
        d= base64.decodestring(fileOrString);
        ex = ExcelFile(BytesIO(d) );
        df = ex.parse(ex.sheet_names[-1])
        df = df.fillna('');
        return df;
        
    if (fileOrString.find("\n") >=0 ):
        ps = [line.strip() for line in fileOrString.split('\n')
                if line.strip() != '' and not line.startswith("#") ];
        if (seperator is None):
            sep = DetermineSeperator(ps[0]);
        else:
            sep = seperator;
        ns = [p.split(sep) for p in ps]
        df1 = pd.DataFrame(ns[1:], columns=ns[0], **kwargs);
    else:               
        df1 = getDF(fileOrString, debug=False, headers=headers, names=names,
                    checkForDateTime=checkForDateTime, 
                    usecols=usecols, seperator=seperator, index_col=index_col, 
                    sheetname=sheetname, xmlTag=xmlTag, proxies=proxies)     

    if ( df1 is None or str(type(df1)).find("DataFrame") < 0):
        return df1;
    #df1=df1.convert_objects(convert_numeric=False)
    df2 = df1[columns] if (columns != None ) else df1;

    if (checkForDateTime):
        for i, c in enumerate(df1.columns):
            if (df2.dtypes[i] != object ):
                continue;
            s = df2[c][0]
            if ( len(s) < 8): continue;
            try:
                dateutil.parser.parse(s);
            except:
                continue; 
            print ("Trying to convert to datetime:"+ c);
            df2[c] =  pd.to_datetime(df2[c])  
            
    if debug:
        print ("Printing 5 of %d rows", df2.shape[0]);
        print (df2[:5]);
        
    return df2;


def normalizeData(df):
    df1 = df.select_dtypes(exclude=[object])
    vals = df1.values
    cols = df1.columns

    d = preprocessing.scale(vals)
    df2  = pd.DataFrame(d, columns=cols)
    return df2;


